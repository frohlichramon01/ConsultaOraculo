import datetime
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options, service=Service(ChromeDriverManager().install()))


### Início - Credenciais para login ###
credenciais = []
with open("login.txt", "r") as file:
    linhas = file.readlines()

for linha in linhas:
    credenciais.append(linha.strip())
usuario = credenciais[0]
senha = credenciais[1]
### Final - Credenciais para login ###


### Início - Cria lista com os nomes a serem consultados ###
lista_tratada = []
with open("nomes.txt", "r", encoding="utf-8") as file:
    linhas = file.readlines()

for _ in linhas:
    lista_tratada.append(_.strip())
### Final - Cria lista com os nomes a serem consultados ###

lista_nomes_final = []
lista_resultado_final = []
lista_dh = []

def abreOraculo():
    driver.get("https://portal.tjpr.jus.br/oraculo/pesquisaReu.do?actionType=start")
    driver.maximize_window()

def fazLogin():
    driver.find_element(By.ID, "username").send_keys(usuario)
    driver.find_element(By.ID, "password").send_keys(senha)
    driver.find_element(By.NAME, "login").click()
    time.sleep(1)

def botaoLimpar():
    driver.find_element(By.NAME, "botaoLimpar").click()

def consulta(nome):
    driver.find_element(By.NAME, "nome").send_keys(nome)
    driver.find_element(By.NAME, "numProcesso").send_keys("000 - Consulta Automatizada Oráculo")
    driver.find_element(By.NAME, "botaoSubmit").click()
    time.sleep(0.3)

def registraDados():
    lista_nomes_final.append(driver.find_element(By.NAME, "nome").get_attribute("value"))
    resultado = driver.find_element(By.CLASS_NAME, "resultTable").text
    if (resultado == "Nenhum item foi encontrado"):
        lista_resultado_final.append("Limpo")
    else:
        lista_resultado_final.append("Verificar")
    lista_dh.append(datetime.datetime.now().strftime("%d-%m-%Y %H:%M"))

def iteraNomes():
    for _ in lista_tratada:
        consulta(_)
        registraDados()
        botaoLimpar()

def criaPlanilha():
    index = 2
    planilha = openpyxl.Workbook()
    consulta_oraculo = planilha["Sheet"]
    consulta_oraculo.title = "Consulta Oráculo"
    consulta_oraculo["A1"] = "Nome Consultado"
    consulta_oraculo["B1"] = "Resultado"
    consulta_oraculo["C1"] = "Data/Hora"

    for nome, result, dh in zip(lista_nomes_final, lista_resultado_final, lista_dh):
        consulta_oraculo.cell(column=1, row=index, value=nome)
        consulta_oraculo.cell(column=2, row=index, value=result)
        consulta_oraculo.cell(column=3, row=index, value=dh)
        index += 1
    
    nome_planilha = "Consulta_Oraculo " + datetime.datetime.now().strftime("%d-%m-%Y %Hh%Mmin") +".xlsx"
    planilha.save(nome_planilha)

def fechaNavegador():
    driver.close()
    

if __name__ == "__main__":
    abreOraculo()
    fazLogin()
    botaoLimpar()
    iteraNomes()
    criaPlanilha()
    fechaNavegador()